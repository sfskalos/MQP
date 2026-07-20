from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path

import numpy as np
import torch
from openpyxl import load_workbook
from transformers import AutoModel, AutoTokenizer


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "data" / "reports.xlsx"
DEFAULT_OUTPUT = ROOT / "data" / "features" / "L" / "biobert_cls_english.npz"
DEFAULT_MODEL = "dmis-lab/biobert-base-cased-v1.1"


def load_reports(workbook_path: Path) -> dict[str, str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    try:
        id_column = headers.index("case_id") + 1
        text_column = headers.index("english_report") + 1
    except ValueError as error:
        raise ValueError(
            "The first worksheet must contain case_id and english_report columns"
        ) from error

    reports: dict[str, str] = {}
    for row in range(2, sheet.max_row + 1):
        case_id = str(sheet.cell(row, id_column).value or "").strip()
        text = str(sheet.cell(row, text_column).value or "").strip()
        if not case_id:
            continue
        if not text:
            raise ValueError(f"Empty report for {case_id}")
        if any("\u4e00" <= character <= "\u9fff" for character in text):
            raise ValueError(f"Non-English report detected for {case_id}")
        reports[case_id] = text
    if len(reports) < 2:
        raise ValueError(f"Expected at least two reports, found {len(reports)}")
    return reports


@torch.inference_mode()
def encode_reports(
    reports: dict[str, str],
    model_name: str,
    max_length: int,
    batch_size: int,
    device: torch.device,
) -> tuple[list[str], np.ndarray]:
    tokenizer = AutoTokenizer.from_pretrained(model_name)
    model = AutoModel.from_pretrained(model_name).to(device).eval()
    names = sorted(reports)
    encoded = []
    for start in range(0, len(names), batch_size):
        batch_names = names[start : start + batch_size]
        inputs = tokenizer(
            [reports[name] for name in batch_names],
            padding=True,
            truncation=True,
            max_length=max_length,
            return_tensors="pt",
        ).to(device)
        output = model(**inputs).last_hidden_state[:, 0]
        encoded.append(output.float().cpu().numpy())
        print(
            json.dumps(
                {"encoded": len(batch_names), "completed": min(start + batch_size, len(names))},
                ensure_ascii=False,
            ),
            flush=True,
        )
    return names, np.concatenate(encoded, axis=0).astype(np.float32)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--max-length", type=int, default=256)
    parser.add_argument("--batch-size", type=int, default=8)
    parser.add_argument("--cpu", action="store_true")
    args = parser.parse_args()

    device = torch.device("cuda" if torch.cuda.is_available() and not args.cpu else "cpu")
    reports = load_reports(args.workbook)
    names, features = encode_reports(
        reports,
        args.model,
        args.max_length,
        args.batch_size,
        device,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    np.savez_compressed(
        args.output,
        case_ids=np.asarray(names),
        features=features,
    )
    manifest = {
        "model_name": "MQP-English-BioBERT",
        "workbook": str(args.workbook),
        "model": args.model,
        "max_length": args.max_length,
        "n": len(names),
        "feature_shape": list(features.shape),
        "device": str(device),
        "sha256": sha256(args.output),
        "output": str(args.output),
    }
    args.output.with_suffix(".json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
